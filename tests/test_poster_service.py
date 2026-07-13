from __future__ import annotations

from pathlib import Path
import tempfile
import unittest
from zipfile import ZipFile

from wp_auto_poster_gui.core.models import Post, PostResult, PosterOptions, UploadedMedia, WordPressConfig
from wp_auto_poster_gui.core.poster_service import export_links_to_source_excel
from wp_auto_poster_gui.core.poster_service import list_website_posts
from wp_auto_poster_gui.core.poster_service import publish_existing_posts_bulk
from wp_auto_poster_gui.core.poster_service import publish_posts
from wp_auto_poster_gui.core.poster_service import publish_from_excel
from wp_auto_poster_gui.core.poster_service import update_existing_posts_image_layout
from wp_auto_poster_gui.core.poster_service import publish_website_posts_bulk
from wp_auto_poster_gui.core.poster_service import update_website_posts_from_excel
from wp_auto_poster_gui.core.poster_service import update_website_posts_image_layout


class FakeClient:
    def __init__(self):
        self.created = []
        self.updated = []
        self.synced = []

    def find_post_by_title(self, title: str):
        if title == "Duplicate":
            return {"id": 99, "link": "https://example.com/duplicate"}
        return None

    def upload_media_from_path(self, path: Path) -> UploadedMedia:
        return UploadedMedia(100 + len(str(path)), f"https://cdn.example.com/{path.name}", path.name)

    def upload_media_from_url(self, url: str) -> UploadedMedia:
        return UploadedMedia(55, url, "remote.jpg")

    def create_post(self, post: Post, content: str, featured_media_id: int | None = None):
        self.created.append((post, content, featured_media_id))
        return {"id": post.row_number, "link": f"https://example.com/{post.row_number}"}

    def update_post(self, post_id: int, post: Post, content: str, featured_media_id: int | None = None):
        self.updated.append((post_id, post, content, featured_media_id))
        return {"link": f"https://example.com/existing/{post_id}"}

    def sync_rank_math_meta(self, post_id: int, post: Post):
        self.synced.append((post_id, list(post.focus_keywords)))
        return {"post_id": post_id, "keyword_count": len(post.focus_keywords)}


class PosterServiceTest(unittest.TestCase):
    def test_publishes_with_local_images_and_updates_duplicate_seo(self) -> None:
        fake = FakeClient()
        posts = [
            Post(2, "First", "<p>A</p><p>B</p>", ma_bai="bai01"),
            Post(3, "Duplicate", "C", ma_bai="bai02", slug="duplicate", focus_keywords=["keyword"]),
        ]
        config = WordPressConfig("https://example.com", "u", "p")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            (folder / "bai01_bg.jpg").write_bytes(b"x")
            (folder / "bai01_1.jpg").write_bytes(b"x")
            (folder / "bai01_2.jpg").write_bytes(b"x")
            (folder / "bai02_bg.jpg").write_bytes(b"x")
            (folder / "bai02_1.jpg").write_bytes(b"x")
            results, orphans = publish_posts(
                posts,
                config,
                PosterOptions(),
                image_folder=folder,
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success", "success"])
        self.assertEqual(orphans, [])
        self.assertEqual(len(fake.created), 1)
        self.assertEqual(len(fake.updated), 1)
        self.assertEqual(fake.synced, [(2, []), (99, ["keyword"])])
        self.assertEqual(fake.updated[0][0], 99)
        self.assertIn("<img", fake.updated[0][2])
        self.assertIsNotNone(fake.updated[0][3])
        self.assertIn("<img", fake.created[0][1])
        self.assertIsNotNone(fake.created[0][2])
        created_content = fake.created[0][1]
        self.assertEqual(created_content.count("<img"), 3)
        self.assertLess(created_content.index("bai01_bg.jpg"), created_content.index("bai01_1.jpg"))
        self.assertLess(created_content.index("bai01_1.jpg"), created_content.index("bai01_2.jpg"))

    def test_publish_from_excel_auto_detects_adjacent_image_zip(self) -> None:
        try:
            import pandas as pd
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeClient()
        config = WordPressConfig("https://example.com", "u", "p")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            excel_path = folder / "posts.xlsx"
            pd.DataFrame(
                [
                    {
                        "ma_bai": "bai01",
                        "Tiêu đề SEO": "First",
                        "Nội dung HTML thuần": "<p>A</p><p>B</p>",
                        "Slug": "first",
                        "Mô tả Meta SEO": "Meta",
                        "Từ khóa chính": "keyword",
                    }
                ]
            ).to_excel(excel_path, sheet_name="Bài SEO HTML", index=False)

            zip_path = folder / "VER2 (2).zip"
            with ZipFile(zip_path, "w") as archive:
                archive.writestr("VER2/bai01_bg.jpg", b"x")
                archive.writestr("VER2/bai01_1.jpg", b"x")

            results, orphans = publish_from_excel(
                excel_path,
                None,
                config,
                PosterOptions(),
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(orphans, [])
        self.assertEqual(len(fake.created), 1)
        self.assertIn("<img", fake.created[0][1])
        self.assertIsNotNone(fake.created[0][2])

    def test_updates_existing_post_image_layout_without_republishing_fields(self) -> None:
        class ImageLayoutClient:
            def __init__(self):
                self.updated_content: tuple[int, str] | None = None

            def find_post_by_slug(self, slug: str):
                if slug.strip("/") == "first":
                    return {"id": 7, "link": "https://example.com/first"}
                return None

            def get_post(self, post_id: int):
                return {
                    "id": post_id,
                    "link": "https://example.com/first",
                    "content": {
                        "raw": '<p>A</p><p><img src="https://example.com/a.jpg" class="wp-image-7 aligncenter" width="300" /></p>'
                    },
                }

            def update_post_content(self, post_id: int, content: str):
                self.updated_content = (post_id, content)
                return {"id": post_id, "link": "https://example.com/first"}

        fake = ImageLayoutClient()

        results = update_existing_posts_image_layout(
            [Post(2, "First", "<p>Excel content is not used here</p>", slug="/first/")],
            WordPressConfig("https://example.com", "u", "p"),
            PosterOptions(image_alignment="alignleft", image_display_size="large"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertIsNotNone(fake.updated_content)
        self.assertEqual(fake.updated_content[0], 7)
        self.assertIn("alignleft", fake.updated_content[1])
        self.assertIn('width="900"', fake.updated_content[1])
        self.assertNotIn("aligncenter", fake.updated_content[1])

    def test_publish_from_excel_force_status_publish(self) -> None:
        try:
            import pandas as pd
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeClient()

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            excel_path = folder / "posts.xlsx"
            pd.DataFrame(
                [
                    {
                        "ma_bai": "bai01",
                        "Tiêu đề SEO": "First",
                        "Nội dung HTML thuần": "<p>A</p>",
                        "Slug": "first",
                        "status": "draft",
                    }
                ]
            ).to_excel(excel_path, sheet_name="Bài SEO HTML", index=False)

            publish_from_excel(
                excel_path,
                None,
                WordPressConfig("https://example.com", "u", "p"),
                PosterOptions(default_status="publish", force_status="publish"),
                client_factory=lambda _: fake,
            )

        self.assertEqual(fake.created[0][0].status, "publish")

    def test_bulk_publish_existing_posts_updates_status_only(self) -> None:
        class BulkClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def find_post_by_slug(self, slug: str):
                if slug.strip("/") == "first":
                    return {"id": 7, "link": "https://example.com/first"}
                return None

            def update_post_status(self, post_id: int, status: str = "publish"):
                self.updated.append((post_id, status))
                return {"id": post_id, "link": "https://example.com/first"}

        fake = BulkClient()
        results = publish_existing_posts_bulk(
            [Post(2, "First", "<p>A</p>", slug="/first/")],
            WordPressConfig("https://example.com", "u", "p"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated, [(7, "publish")])

    def test_list_website_posts_uses_client_list_posts(self) -> None:
        class WebsiteListClient:
            def list_posts(self):
                return [
                    {"id": 9, "title": {"rendered": "Bài web"}, "status": "draft"},
                ]

        messages: list[str] = []
        posts = list_website_posts(
            WordPressConfig("https://example.com", "u", "p"),
            progress_callback=messages.append,
            client_factory=lambda _: WebsiteListClient(),
        )

        self.assertEqual(posts[0]["id"], 9)
        self.assertIn("Đã tải 1 bài viết", messages[-1])

    def test_publish_website_posts_bulk_updates_selected_ids(self) -> None:
        class WebsitePublishClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def update_post_status(self, post_id: int, status: str = "publish"):
                self.updated.append((post_id, status))
                return {"id": post_id, "link": f"https://example.com/{post_id}"}

        fake = WebsitePublishClient()
        results = publish_website_posts_bulk(
            [{"id": 15, "title": {"rendered": "Bài nháp"}, "link": "https://example.com/draft"}],
            WordPressConfig("https://example.com", "u", "p"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated, [(15, "publish")])
        self.assertEqual(results[0].row_number, 15)

    def test_update_website_posts_image_layout_updates_content_by_id(self) -> None:
        class WebsiteImageClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def get_post(self, post_id: int):
                return {
                    "id": post_id,
                    "title": {"rendered": "Bài có ảnh"},
                    "link": "https://example.com/with-image",
                    "content": {"raw": '<p><img src="https://example.com/a.jpg" /></p>'},
                }

            def update_post_content(self, post_id: int, content: str):
                self.updated.append((post_id, content))
                return {"id": post_id, "link": "https://example.com/with-image"}

        fake = WebsiteImageClient()
        results = update_website_posts_image_layout(
            [{"id": 21, "title": {"rendered": "Bài có ảnh"}}],
            WordPressConfig("https://example.com", "u", "p"),
            PosterOptions(image_alignment="aligncenter", image_display_size="medium", image_custom_width=600),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated[0][0], 21)
        self.assertIn("aligncenter", fake.updated[0][1])
        self.assertIn("width=\"600\"", fake.updated[0][1])

    def test_add_images_to_website_posts_matches_slug_and_skips_existing_image(self) -> None:
        from wp_auto_poster_gui.core.poster_service import add_images_to_website_posts

        class WebsiteAddImagesClient:
            def __init__(self):
                self.uploaded: list[str] = []
                self.updated: list[tuple[int, str]] = []

            def get_post(self, post_id: int):
                return {
                    "id": post_id,
                    "title": {"rendered": "Bài có ảnh"},
                    "link": "https://example.com/bai-co-anh",
                    "content": {
                        "raw": '<p>A</p><p><img src="https://example.com/uploads/bai-co-anh-1.jpg" '
                        'data-pmedia-source="bai-co-anh_1.jpg" /></p><p>B</p>'
                    },
                }

            def upload_media_from_path(self, path: Path):
                self.uploaded.append(path.name)
                return UploadedMedia(
                    100 + len(self.uploaded),
                    f"https://example.com/uploads/{path.name}",
                    path.name,
                )

            def update_post_content(self, post_id: int, content: str):
                self.updated.append((post_id, content))
                return {"id": post_id, "link": "https://example.com/bai-co-anh"}

        fake = WebsiteAddImagesClient()
        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            (folder / "bai-co-anh_thumb.jpg").write_bytes(b"thumb")
            (folder / "bai-co-anh_1.jpg").write_bytes(b"existing")
            (folder / "bai-co-anh_2.jpg").write_bytes(b"new")
            results, orphan_files = add_images_to_website_posts(
                [{"id": 21, "slug": "bai-co-anh", "title": {"rendered": "Bài có ảnh"}}],
                folder,
                WordPressConfig("https://example.com", "u", "p"),
                PosterOptions(max_images_per_post=3, image_alignment="aligncenter"),
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.uploaded, ["bai-co-anh_thumb.jpg", "bai-co-anh_2.jpg"])
        self.assertEqual(fake.updated[0][1].count("<img"), 3)
        self.assertIn('data-pmedia-source="bai-co-anh_thumb.jpg"', fake.updated[0][1])
        self.assertEqual(orphan_files, [])

    def test_export_links_to_source_excel_appends_adjacent_link_column(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            source_path = folder / "posts.xlsx"
            output_path = folder / "posts_with_links.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Bài SEO HTML"
            worksheet.append(["Tiêu đề SEO", "Nội dung HTML thuần", "Slug"])
            worksheet.append(["Bài 1", "<p>A</p>", "bai-1"])
            worksheet.append(["Bài 2", "<p>B</p>", "bai-2"])
            workbook.save(source_path)

            exported = export_links_to_source_excel(
                source_path,
                [
                    PostResult(2, "Bài 1", "success", link="https://example.com/bai-1"),
                    PostResult(3, "Bài 2", "failed", error="Lỗi đăng"),
                ],
                output_path,
            )

            exported_workbook = load_workbook(exported)
            exported_sheet = exported_workbook["Bài SEO HTML"]

            self.assertEqual(exported, output_path)
            self.assertEqual(exported_sheet.cell(row=1, column=4).value, "Link bài viết")
            self.assertEqual(exported_sheet.cell(row=2, column=4).value, "https://example.com/bai-1")
            self.assertEqual(exported_sheet.cell(row=2, column=4).hyperlink.target, "https://example.com/bai-1")
            self.assertIsNone(exported_sheet.cell(row=3, column=4).value)

            original_workbook = load_workbook(source_path)
            self.assertEqual(original_workbook["Bài SEO HTML"].max_column, 3)

    def test_export_links_to_source_excel_reuses_existing_link_column(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            source_path = folder / "posts.xlsx"
            output_path = folder / "posts_with_links.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Bài SEO HTML"
            worksheet.append(["Tiêu đề SEO", "Nội dung HTML thuần", "Link bài viết"])
            worksheet.append(["Bài 1", "<p>A</p>", ""])
            workbook.save(source_path)

            export_links_to_source_excel(
                source_path,
                [PostResult(2, "Bài 1", "success", link="https://example.com/bai-1")],
                output_path,
            )

            exported_sheet = load_workbook(output_path)["Bài SEO HTML"]
            self.assertEqual(exported_sheet.max_column, 3)
            self.assertEqual(exported_sheet.cell(row=2, column=3).value, "https://example.com/bai-1")


class FakeSiteClient:
    """Fake WordPress client with three posts in different completeness states."""

    def __init__(self):
        self.posts = {
            101: {
                "id": 101,
                "link": "https://example.com/bai-mot",
                "slug": "bai-mot",
                "content": {"raw": "<p>Đã có nội dung</p>"},
                "excerpt": {"raw": ""},
                "categories": [2],
                "tags": [5],
                "featured_media": 9,
                "meta": {
                    "rank_math_title": "Tiêu đề SEO cũ",
                    "rank_math_description": "",
                    "rank_math_focus_keyword": "old keyword",
                },
            },
            102: {
                "id": 102,
                "link": "https://example.com/bai-hai",
                "slug": "bai-hai",
                "content": {"raw": "<p>Nội dung</p>"},
                "excerpt": {"raw": "Mô tả sẵn"},
                "categories": [1],
                "tags": [],
                "featured_media": 0,
                "meta": {
                    "rank_math_title": "SEO",
                    "rank_math_description": "Mô tả",
                    "rank_math_focus_keyword": "kw",
                },
            },
            103: {
                "id": 103,
                "link": "https://example.com/bai-ba",
                "slug": "bai-ba",
                "content": {"raw": "<p>Nội dung</p>"},
                "excerpt": {"raw": "Mô tả đầy đủ"},
                "categories": [3],
                "tags": [7],
                "featured_media": 12,
                "meta": {
                    "rank_math_title": "SEO",
                    "rank_math_description": "Mô tả",
                    "rank_math_focus_keyword": "kw",
                },
            },
        }
        self.updated_fields: list[tuple[int, dict]] = []
        self.synced: list[tuple[int, dict]] = []
        self.terms: list[tuple[str, str]] = []

    def get_post(self, post_id: int):
        if post_id not in self.posts:
            raise RuntimeError(f"Post {post_id} not found")
        return self.posts[post_id]

    def find_post_by_slug(self, slug: str):
        normalized = slug.strip().split("?", 1)[0].split("#", 1)[0]
        if "://" in normalized:
            normalized = normalized.rstrip("/").rsplit("/", 1)[-1]
        normalized = normalized.strip("/")
        for post in self.posts.values():
            if post.get("slug") == normalized:
                return post
        return None

    def find_post_by_title(self, title: str):
        return None

    def get_or_create_term(self, taxonomy: str, name: str) -> int:
        self.terms.append((taxonomy, name))
        return 77 if taxonomy == "categories" else 88

    def upload_media_from_url(self, url: str) -> UploadedMedia:
        return UploadedMedia(55, url, "anh.jpg")

    def update_post_fields(self, post_id: int, fields: dict):
        self.updated_fields.append((post_id, fields))
        return {"id": post_id, "link": self.posts[post_id]["link"]}

    def sync_rank_math_fields(self, post_id: int, fields: dict):
        self.synced.append((post_id, fields))
        return {"supported": True, "post_id": post_id}


class UpdateWebsitePostsFromExcelTest(unittest.TestCase):
    def _write_workbook(self, directory: str, rows: list[dict]) -> Path:
        import pandas as pd

        workbook = Path(directory) / "updates.xlsx"
        pd.DataFrame(rows).to_excel(workbook, sheet_name="Cập nhật", index=False)
        return workbook

    def test_fill_missing_only_patches_empty_fields(self) -> None:
        try:
            import pandas as pd  # noqa: F401
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeSiteClient()
        config = WordPressConfig("https://example.com", "u", "p")
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                directory,
                [
                    {"ID": 101, "Mô tả Meta SEO": "Mô tả mới", "Từ khóa chính": "kw mới"},
                    {
                        "Link bài viết": "https://example.com/bai-hai/",
                        "Danh mục": "Tin Tức",
                        "Tags": "tag mới",
                        "URL ảnh đại diện": "https://cdn.example.com/anh.jpg",
                    },
                    {"Tiêu đề SEO": "Không tồn tại", "Mô tả Meta SEO": "abc"},
                    {"ID": 103, "Mô tả Meta SEO": "Mô tả khác"},
                ],
            )
            results = update_website_posts_from_excel(
                workbook,
                config,
                client_factory=lambda _: fake,
            )

        self.assertEqual(
            [result.status for result in results],
            ["success", "success", "skipped", "skipped"],
        )

        # Post 101: existing keywords are kept, missing description/excerpt are filled.
        post_101_fields = dict(fake.updated_fields)[101]
        self.assertEqual(post_101_fields["excerpt"], "Mô tả mới")
        self.assertEqual(post_101_fields["meta"], {"rank_math_description": "Mô tả mới"})
        self.assertNotIn("content", post_101_fields)
        post_101_sync = dict(fake.synced)[101]
        self.assertEqual(post_101_sync["focus_keywords"], ["old keyword"])
        self.assertEqual(post_101_sync["meta_description"], "Mô tả mới")
        self.assertNotIn("seo_title", post_101_sync)

        # Post 102: default category, empty tags and missing featured image are filled.
        post_102_fields = dict(fake.updated_fields)[102]
        self.assertEqual(post_102_fields["categories"], [77])
        self.assertEqual(post_102_fields["tags"], [88])
        self.assertEqual(post_102_fields["featured_media"], 55)
        self.assertNotIn("meta", post_102_fields)
        self.assertNotIn(102, dict(fake.synced))
        self.assertIn(("categories", "Tin Tức"), fake.terms)
        self.assertIn(("tags", "tag mới"), fake.terms)

        # Row 3 has no matching post; row 4 already has every provided field.
        self.assertIn("Không tìm thấy", results[2].error)
        self.assertIn("đủ dữ liệu", results[3].error)
        self.assertEqual({post_id for post_id, _ in fake.updated_fields}, {101, 102})

    def test_overwrite_mode_replaces_existing_values(self) -> None:
        try:
            import pandas as pd  # noqa: F401
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeSiteClient()
        config = WordPressConfig("https://example.com", "u", "p")
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                directory,
                [
                    {
                        "ID": 101,
                        "Từ khóa chính": "kw mới",
                        "Trạng thái": "draft",
                        "Nội dung HTML": "<p>Nội dung mới</p>",
                    }
                ],
            )
            results = update_website_posts_from_excel(
                workbook,
                config,
                only_fill_missing=False,
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success"])
        post_101_fields = dict(fake.updated_fields)[101]
        self.assertEqual(post_101_fields["status"], "draft")
        self.assertEqual(post_101_fields["content"], "<p>Nội dung mới</p>")
        self.assertEqual(post_101_fields["meta"]["rank_math_focus_keyword"], "kw mới")
        post_101_sync = dict(fake.synced)[101]
        self.assertEqual(post_101_sync["focus_keywords"], ["kw mới"])

    def test_stop_event_halts_processing(self) -> None:
        try:
            import pandas as pd  # noqa: F401
        except ImportError:
            self.skipTest("pandas is not installed")

        from threading import Event

        fake = FakeSiteClient()
        config = WordPressConfig("https://example.com", "u", "p")
        stop_event = Event()
        stop_event.set()
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                directory,
                [{"ID": 101, "Mô tả Meta SEO": "Mô tả mới"}],
            )
            results = update_website_posts_from_excel(
                workbook,
                config,
                stop_event=stop_event,
                client_factory=lambda _: fake,
            )

        self.assertEqual(results, [])
        self.assertEqual(fake.updated_fields, [])


if __name__ == "__main__":
    unittest.main()
