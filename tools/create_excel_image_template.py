from __future__ import annotations

from pathlib import Path


OUTPUT_XLSX = Path("docs/Mau_Excel_Dang_Bai_SEO_HTML_WordPress.xlsx")
OUTPUT_IMAGES = Path("docs/mau_ten_hinh_anh")
OUTPUT_NAMING_TXT = Path("docs/Quy_uoc_dat_ten_hinh_anh.txt")


def main() -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from PIL import Image, ImageDraw, ImageFont
    except ImportError as exc:
        print(f"Missing dependency: {exc}")
        return 1

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_IMAGES.mkdir(parents=True, exist_ok=True)
    for old_image in OUTPUT_IMAGES.iterdir():
        if old_image.is_file() and old_image.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            old_image.unlink()

    workbook = Workbook()
    posts_sheet = workbook.active
    posts_sheet.title = "Bài SEO HTML"

    post_headers = [
        "STT",
        "ma_bai",
        "Từ khóa chính",
        "Từ khóa phụ đã phủ thêm",
        "Tiêu đề SEO",
        "Nội dung HTML thuần",
        "Slug",
        "Mô tả Meta SEO",
        "Danh mục",
        "tags",
        "status",
        "publish_date",
        "featured_image_url",
    ]
    post_rows = [
        [
            1,
            "thong-cong-bien-hoa",
            "thông cống Biên Hòa",
            "thông cống Đồng Nai, thông cống nghẹt, xử lý cống nghẹt",
            "Thông Cống Biên Hòa Nhanh Gọn, Có Mặt Kịp Thời",
            (
                "<h2>Thông Cống Biên Hòa Nhanh Gọn, Có Mặt Kịp Thời</h2>\n"
                "<p><strong>thông cống Biên Hòa</strong> là nhu cầu thường gặp khi đường ống thoát nước bị nghẹt, "
                "rút chậm hoặc phát sinh mùi hôi trong sinh hoạt.</p>\n"
                "<h3>Khi nào nên gọi thợ thông cống?</h3>\n"
                "<p>Nên gọi thợ khi nước rút chậm, bồn rửa trào ngược hoặc đã thử xử lý nhưng tình trạng vẫn lặp lại.</p>\n"
                "<h3>Quy trình xử lý</h3>\n"
                "<p>Thợ kiểm tra vị trí nghẹt, báo giá rõ ràng, xử lý bằng thiết bị phù hợp và vệ sinh lại khu vực làm việc.</p>\n"
                "<p>Liên hệ để được hỗ trợ nhanh tại Biên Hòa và khu vực lân cận.</p>"
            ),
            "thong-cong-bien-hoa",
            "Dịch vụ thông cống Biên Hòa hỗ trợ nhanh, xử lý cống nghẹt, nước rút chậm và mùi hôi với báo giá rõ ràng.",
            "Dịch vụ",
            "thông cống Biên Hòa, thông cống Đồng Nai",
            "draft",
            "",
            "",
        ],
        [
            2,
            "thong-cong-thu-duc",
            "thông cống Thủ Đức",
            "thông cống gần đây, thợ thông cống, thông cống 24h",
            "Thông Cống Thủ Đức Hỗ Trợ Nhanh Khi Cống Nghẹt Và Nước Trào",
            (
                "<h2>Thông Cống Thủ Đức Hỗ Trợ Nhanh Khi Cống Nghẹt Và Nước Trào</h2>\n"
                "<p><strong>thông cống Thủ Đức</strong> giúp xử lý các sự cố nghẹt đường ống trong nhà ở, quán ăn, "
                "văn phòng hoặc nhà trọ.</p>\n"
                "<h3>Dấu hiệu cần xử lý sớm</h3>\n"
                "<p>Nước rút chậm, mùi hôi quay lại, lavabo hoặc hố ga trào nước là các dấu hiệu cần kiểm tra.</p>\n"
                "<h3>Lưu ý trước khi đặt lịch</h3>\n"
                "<p>Chuẩn bị mô tả vị trí nghẹt, hình ảnh hiện trạng và thời gian có thể cho thợ kiểm tra.</p>"
            ),
            "thong-cong-thu-duc",
            "Thông cống Thủ Đức hỗ trợ xử lý cống nghẹt, hố ga trào, bồn rửa chén thoát chậm và mùi hôi.",
            "Dịch vụ",
            "thông cống Thủ Đức, thợ thông cống",
            "draft",
            "",
            "",
        ],
        [
            3,
            "rut-ham-cau-dong-nai",
            "rút hầm cầu Đồng Nai",
            "hút hầm cầu, rút hầm cầu gần đây, xử lý hầm cầu đầy",
            "Rút Hầm Cầu Đồng Nai Đúng Quy Trình, Hạn Chế Mùi Hôi",
            (
                "<h2>Rút Hầm Cầu Đồng Nai Đúng Quy Trình, Hạn Chế Mùi Hôi</h2>\n"
                "<p><strong>rút hầm cầu Đồng Nai</strong> cần thực hiện đúng quy trình để hạn chế mùi hôi, tràn ngược "
                "và ảnh hưởng sinh hoạt.</p>\n"
                "<h3>Khi nào hầm cầu có thể đã đầy?</h3>\n"
                "<p>Bồn cầu xả yếu, có mùi nặng hoặc nước thoát chậm nhiều ngày là các dấu hiệu nên kiểm tra.</p>\n"
                "<h3>Thông tin cần cung cấp</h3>\n"
                "<p>Địa chỉ, vị trí nắp hầm, tình trạng hiện tại và khung giờ có mặt để đội kỹ thuật khảo sát.</p>"
            ),
            "rut-ham-cau-dong-nai",
            "Rút hầm cầu Đồng Nai hỗ trợ khảo sát, báo giá rõ ràng và xử lý đúng quy trình cho nhà ở, công ty, quán ăn.",
            "Dịch vụ",
            "rút hầm cầu Đồng Nai, hút hầm cầu",
            "draft",
            "",
            "",
        ],
    ]

    write_sheet(posts_sheet, post_headers, post_rows)
    add_post_comments(posts_sheet)
    add_status_validation(posts_sheet)

    naming_sheet = workbook.create_sheet("Quy ước đặt tên ảnh")
    naming_headers = ["Mẫu tên file", "Vai trò", "Ví dụ", "Ghi chú"]
    naming_rows = [
        ["{ma_bai}_bg.jpg", "Ảnh đại diện/thumb", "thong-cong-bien-hoa_bg.jpg", "Được upload làm featured image và chèn là hình đầu tiên trong bài."],
        ["{ma_bai}_thumb.webp", "Ảnh đại diện/thumb", "thong-cong-bien-hoa_thumb.webp", "Có thể dùng thumb, thumbnail hoặc featured thay cho bg."],
        ["{ma_bai}_1.jpg", "Ảnh nội dung 1", "thong-cong-bien-hoa_1.jpg", "Được upload và chèn sau ảnh thumb trong bài."],
        ["{ma_bai}_2.png", "Ảnh nội dung 2", "thong-cong-bien-hoa_2.png", "Đánh số tăng dần: _1, _2, _3..."],
        ["{ma_bai}_3.webp", "Ảnh nội dung tiếp theo", "thong-cong-bien-hoa_3.webp", "App chỉ chèn tối đa theo số ảnh đã chọn trong giao diện."],
        ["anh-khong-khop.jpg", "Không khớp bài nào", "anh-khong-khop.jpg", "App sẽ báo trong preview để kiểm tra lại tên file."],
    ]
    write_sheet(naming_sheet, naming_headers, naming_rows)

    checklist_sheet = workbook.create_sheet("Checklist trước khi đăng")
    checklist_headers = ["Mục kiểm tra", "Đạt khi", "Ghi chú"]
    checklist_rows = [
        ["Kết nối WordPress", "Nút kiểm tra kết nối báo OK", "URL phải có https:// và dùng Application Password."],
        ["Sheet Bài SEO HTML", "Có cột Tiêu đề SEO và Nội dung HTML thuần", "App ưu tiên sheet có tên Bài SEO HTML."],
        ["ma_bai", "Mỗi bài có mã ngắn, không dấu, không khoảng trắng", "Mã này dùng để nhận diện ảnh local."],
        ["Nội dung HTML", "Có H2/H3/P và phần liên hệ nếu cần", "Không để trống title hoặc content."],
        ["Slug", "Không dấu, dùng dấu gạch ngang", "Ví dụ: thong-cong-bien-hoa."],
        ["Meta SEO", "Mô tả dưới khoảng 160 ký tự nếu có thể", "App gửi vào excerpt và Rank Math description."],
        ["Từ khóa", "Từ khóa chính và phụ tách rõ bằng dấu phẩy", "App đưa vào Rank Math focus keyword."],
        ["Ảnh đại diện", "Có file {ma_bai}_bg hoặc {ma_bai}_thumb nếu muốn dùng ảnh local", "Ảnh này cũng là hình đầu tiên trong nội dung."],
        ["Ảnh nội dung", "Tên file dạng {ma_bai}_1, {ma_bai}_2", "Đuôi hỗ trợ: jpg, jpeg, png, webp."],
        ["Trạng thái", "Dùng draft trước khi kiểm tra xong", "Chỉ dùng publish khi chắc chắn nội dung đúng."],
        ["Xuất kết quả", "Sau khi đăng có file Excel mới kèm Link bài viết", "Cột Link bài viết được thêm sát bên phải dữ liệu gốc."],
        ["Lịch sử chạy", "Tab Lịch sử có bản ghi lần chạy", "Lưu local trong config/run_history.json."],
    ]
    write_sheet(checklist_sheet, checklist_headers, checklist_rows)

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    workbook.save(OUTPUT_XLSX)
    create_sample_images(Image, ImageDraw, ImageFont)
    write_naming_text()

    print(OUTPUT_XLSX)
    print(OUTPUT_IMAGES)
    print(OUTPUT_NAMING_TXT)
    return 0


def write_sheet(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def style_sheet(sheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="1F4D78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 8,
        "B": 24,
        "C": 28,
        "D": 42,
        "E": 48,
        "F": 72,
        "G": 28,
        "H": 50,
        "I": 18,
        "J": 34,
        "K": 14,
        "L": 22,
        "M": 34,
    }
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = widths.get(letter, min(max(max_length + 2, 16), 52))
    sheet.row_dimensions[1].height = 34


def add_post_comments(sheet) -> None:
    from openpyxl.comments import Comment

    comments = {
        "B1": "Mã bài dùng để nhận diện ảnh. Ví dụ ảnh: thong-cong-bien-hoa_bg.jpg, thong-cong-bien-hoa_1.jpg.",
        "C1": "Từ khóa chính được đưa vào Rank Math focus keyword.",
        "D1": "Từ khóa phụ tách bằng dấu phẩy. App tự gộp với từ khóa chính và bỏ trùng.",
        "E1": "Tiêu đề bài viết WordPress và Rank Math SEO title.",
        "F1": "Nội dung HTML thuần. App giữ HTML và chèn ảnh local vào bài.",
        "G1": "Slug/đường dẫn cố định. Nên không dấu, dùng dấu gạch ngang.",
        "H1": "Đưa vào excerpt WordPress và Rank Math description.",
        "J1": "Thẻ WordPress, tách bằng dấu phẩy. Không bắt buộc.",
        "K1": "Nên dùng draft trước khi kiểm tra xong. Có thể dùng publish hoặc future nếu cần.",
        "L1": "Ngày đăng dạng ISO, ví dụ 2026-07-10T09:00:00. Có thể để trống.",
        "M1": "URL ảnh đại diện từ internet. Nếu có ảnh local _bg/_thumb, app ưu tiên ảnh local.",
    }
    for cell, text in comments.items():
        sheet[cell].comment = Comment(text, "PMEDIA")


def add_status_validation(sheet) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    validation = DataValidation(type="list", formula1='"draft,publish,future,private"', allow_blank=False)
    validation.error = "Chỉ nên dùng draft, publish, future hoặc private."
    validation.errorTitle = "Status không hợp lệ"
    validation.prompt = "Khuyến nghị dùng draft để kiểm tra trước khi public."
    validation.promptTitle = "Trạng thái bài viết"
    sheet.add_data_validation(validation)
    validation.add("K2:K500")


def create_sample_images(Image, ImageDraw, ImageFont) -> None:
    sample_images = [
        ("thong-cong-bien-hoa_bg.jpg", "#1F4D78", "Ảnh đại diện"),
        ("thong-cong-bien-hoa_1.jpg", "#2E74B5", "Ảnh nội dung 1"),
        ("thong-cong-bien-hoa_2.jpg", "#3A8E7A", "Ảnh nội dung 2"),
        ("thong-cong-thu-duc_bg.jpg", "#7A5A00", "Ảnh đại diện"),
        ("thong-cong-thu-duc_1.jpg", "#B9794B", "Ảnh nội dung 1"),
        ("rut-ham-cau-dong-nai_thumb.webp", "#5B5F97", "Ảnh đại diện"),
        ("rut-ham-cau-dong-nai_1.webp", "#F4A261", "Ảnh nội dung 1"),
        ("anh-khong-khop.jpg", "#777777", "File không khớp ma_bai"),
    ]
    for filename, color, label in sample_images:
        image = Image.new("RGB", (1200, 800), color)
        draw = ImageDraw.Draw(image)
        draw.rectangle((36, 36, 1164, 764), outline="white", width=6)
        draw.rectangle((70, 520, 1130, 700), outline="white", width=3)
        draw.text((90, 90), filename, fill="white")
        draw.text((90, 150), label, fill="white")
        draw.text((90, 570), "Tên file mẫu để app nhận diện ảnh thuộc bài nào", fill="white")
        image.save(OUTPUT_IMAGES / filename, quality=88)


def write_naming_text() -> None:
    OUTPUT_NAMING_TXT.write_text(
        "\n".join(
            [
                "QUY ƯỚC ĐẶT TÊN HÌNH ẢNH CHO PMEDIA-ĐĂNG BÀI TỰ ĐỘNG WORDPRESS",
                "",
                "1. Ảnh đại diện/thumb:",
                "   {ma_bai}_bg.jpg",
                "   {ma_bai}_thumb.jpg",
                "   {ma_bai}_thumbnail.png",
                "   {ma_bai}_featured.webp",
                "",
                "2. Ảnh nội dung trong bài:",
                "   {ma_bai}_1.jpg",
                "   {ma_bai}_2.png",
                "   {ma_bai}_3.webp",
                "",
                "3. Ví dụ với ma_bai = thong-cong-bien-hoa:",
                "   thong-cong-bien-hoa_bg.jpg",
                "   thong-cong-bien-hoa_1.jpg",
                "   thong-cong-bien-hoa_2.jpg",
                "",
                "4. Đuôi file được hỗ trợ:",
                "   .jpg, .jpeg, .png, .webp",
                "",
                "5. Lưu ý:",
                "   - ma_bai trong Excel phải trùng phần đầu tên file ảnh.",
                "   - Dùng dấu gạch dưới _ sau ma_bai.",
                "   - Không đặt khoảng trắng hoặc tiếng Việt có dấu trong ma_bai.",
                "   - File không khớp ma_bai sẽ hiện cảnh báo trong preview.",
            ]
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    raise SystemExit(main())
